"""
Export for Accountant - generates PDF + Excel from RigBooks session data.
"""
import pandas as pd
import json
import io
import os
from datetime import datetime
from pathlib import Path

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, HRFlowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _get_phone_data(phone_bill):
    results = {}
    for person in ['greg', 'lilibeth']:
        data = phone_bill.get(person, {})
        if 'months' in data and isinstance(data['months'], dict):
            annual = sum(data['months'].values())
        else:
            annual = data.get('monthly', 0.0) * 12
        biz_pct = data.get('business_pct', 100)
        deductible = annual * biz_pct / 100
        itc = deductible * 0.05 / 1.05
        results[person] = {
            'annual': annual, 'business_pct': biz_pct,
            'deductible': deductible, 'itc': itc,
            'months': data.get('months', {})
        }
    return results


def _build_revenue_breakdown(df):
    if df is None or df.empty:
        return {'wire': pd.DataFrame(), 'mobile': pd.DataFrame(),
                'branch': pd.DataFrame(), 'etransfer': pd.DataFrame(),
                'wire_total': 0, 'mobile_total': 0, 'branch_total': 0,
                'etransfer_total': 0, 'grand_total': 0}
    credits = df[df['credit'] > 0].copy()
    wire_mask = credits['description'].str.contains('WIRE TSF', case=False, na=False)
    mobile_mask = credits['description'].str.contains('MOBILE DEP', case=False, na=False) & ~wire_mask
    branch_mask = credits['description'].str.contains('BRANCH DEP|DEPOSIT', case=False, na=False) & ~wire_mask & ~mobile_mask
    etransfer_mask = credits['description'].str.contains('E-TRANSFER|INTERAC', case=False, na=False) & ~wire_mask & ~mobile_mask & ~branch_mask
    wire_df = credits[wire_mask]
    mobile_df = credits[mobile_mask]
    branch_df = credits[branch_mask]
    etransfer_df = credits[etransfer_mask]
    return {
        'wire': wire_df, 'mobile': mobile_df,
        'branch': branch_df, 'etransfer': etransfer_df,
        'wire_total': wire_df['credit'].sum(),
        'mobile_total': mobile_df['credit'].sum(),
        'branch_total': branch_df['credit'].sum(),
        'etransfer_total': etransfer_df['credit'].sum(),
        'grand_total': credits['credit'].sum()
    }


def _build_expense_breakdown(df):
    if df is None or df.empty:
        return {}
    expenses = df[df['debit'] > 0].copy()
    if 'cra_category' not in expenses.columns:
        return {'Uncategorized': {'total': expenses['debit'].sum(), 'count': len(expenses), 'itc': 0}}
    result = {}
    for cat in expenses['cra_category'].dropna().unique():
        cat_df = expenses[expenses['cra_category'] == cat]
        itc = cat_df['itc_amount'].sum() if 'itc_amount' in cat_df.columns else 0
        result[cat] = {'total': cat_df['debit'].sum(), 'count': len(cat_df), 'itc': itc}
    return result


def generate_pdf(classified_df, cash_expenses, phone_bill, fiscal_year="2024-2025"):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter,
                            topMargin=0.6*inch, bottomMargin=0.6*inch,
                            leftMargin=0.7*inch, rightMargin=0.7*inch)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='SectionHead', parent=styles['Heading2'],
                              spaceAfter=6, spaceBefore=14, textColor=colors.HexColor('#1a1a5e')))
    styles.add(ParagraphStyle(name='SmallNote', parent=styles['Normal'],
                              fontSize=8, textColor=colors.grey))
    story = []
    fy_start, fy_end = fiscal_year.split('-')

    story.append(Paragraph("Cape Bretoner's Oilfield Services Ltd.", styles['Title']))
    story.append(Paragraph(f"Fiscal Year {fiscal_year}  (Dec 1, {fy_start} - Nov 30, {fy_end})", styles['Normal']))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", styles['SmallNote']))
    story.append(Spacer(1, 6))
    story.append(Paragraph("Ownership: Greg MacDonald 51% | Lilibeth Sejera 49%", styles['Normal']))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
    story.append(Spacer(1, 10))

    def make_table(headers, rows, col_widths=None):
        data = [headers] + rows
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a1a5e')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 9),
            ('FONTSIZE', (0,1), (-1,-1), 8),
            ('ALIGN', (-1,0), (-1,-1), 'RIGHT'),
            ('ALIGN', (-2,0), (-2,-1), 'RIGHT'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f5f5ff')]),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cccccc')),
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ]))
        return t

    # 1. Revenue
    story.append(Paragraph("1. Revenue Breakdown", styles['SectionHead']))
    rev = _build_revenue_breakdown(classified_df)
    rev_rows = [
        ['Wire Transfers (Long Run / PWC)', f"${rev['wire_total']:,.2f}"],
        ['Mobile Deposits', f"${rev['mobile_total']:,.2f}"],
        ['Branch Deposits', f"${rev['branch_total']:,.2f}"],
        ['E-Transfers (Interac)', f"${rev['etransfer_total']:,.2f}"],
        ['TOTAL REVENUE', f"${rev['grand_total']:,.2f}"],
    ]
    story.append(make_table(['Source', 'Amount'], rev_rows, [4*inch, 2*inch]))
    story.append(Spacer(1, 4))

    if not rev['wire'].empty:
        story.append(Paragraph("Wire Transfer Detail:", styles['Normal']))
        wire_rows = []
        for _, r in rev['wire'].iterrows():
            d = r.get('date', '')
            if hasattr(d, 'strftime'): d = d.strftime('%Y-%m-%d')
            wire_rows.append([str(d), str(r['description'])[:50], f"${r['credit']:,.2f}"])
        story.append(make_table(['Date', 'Description', 'Amount'], wire_rows, [1.2*inch, 3.5*inch, 1.3*inch]))
        story.append(Spacer(1, 6))

    gst_collected = rev['grand_total'] * 0.05
    story.append(Paragraph(f"GST Collected on Revenue (5%): <b>${gst_collected:,.2f}</b>", styles['Normal']))
    story.append(Spacer(1, 10))

    # 2. Expenses
    story.append(Paragraph("2. Business Expenses by CRA Category", styles['SectionHead']))
    exp = _build_expense_breakdown(classified_df)
    exp_rows = []
    total_exp = 0
    total_bank_itc = 0
    for cat, info in sorted(exp.items()):
        exp_rows.append([cat, str(info['count']), f"${info['total']:,.2f}", f"${info['itc']:,.2f}"])
        total_exp += info['total']
        total_bank_itc += info['itc']
    exp_rows.append(['TOTAL', '', f"${total_exp:,.2f}", f"${total_bank_itc:,.2f}"])
    story.append(make_table(['Category', 'Txns', 'Amount', 'ITC'], exp_rows, [2.5*inch, 0.8*inch, 1.5*inch, 1.2*inch]))
    story.append(Spacer(1, 10))

    # 3. Cash Expenses
    story.append(Paragraph("3. Cash Expenses (Not in Bank Statement)", styles['SectionHead']))
    cash_itc = 0
    if cash_expenses:
        cash_rows = []
        for e in cash_expenses:
            amt = e.get('amount', 0)
            itc = amt * 0.05 / 1.05
            cash_itc += itc
            receipt = "Yes" if e.get('has_receipt', False) else ("Recommended" if amt >= 30 else "Not required")
            cash_rows.append([e.get('date','N/A'), e.get('description','N/A')[:40], e.get('category','N/A'),
                              f"${amt:,.2f}", f"${itc:,.2f}", receipt])
        cash_rows.append(['','','TOTAL', f"${sum(e.get('amount',0) for e in cash_expenses):,.2f}", f"${cash_itc:,.2f}", ''])
        story.append(make_table(['Date','Description','Category','Amount','ITC','Receipt'], cash_rows,
                                [0.9*inch, 1.5*inch, 1*inch, 0.9*inch, 0.8*inch, 0.9*inch]))
    else:
        story.append(Paragraph("No cash expenses recorded.", styles['Normal']))
    story.append(Spacer(1, 10))

    # 4. Phone Bills
    story.append(Paragraph("4. Phone Bill Deductions", styles['SectionHead']))
    phone = _get_phone_data(phone_bill)
    phone_rows = []
    total_phone_itc = 0
    for person, label in [('greg','Greg MacDonald'), ('lilibeth','Lilibeth Sejera')]:
        p = phone[person]
        phone_rows.append([label, f"${p['annual']:,.2f}", f"{p['business_pct']}%", f"${p['deductible']:,.2f}", f"${p['itc']:,.2f}"])
        total_phone_itc += p['itc']
    phone_rows.append(['TOTAL','','', f"${phone['greg']['deductible']+phone['lilibeth']['deductible']:,.2f}", f"${total_phone_itc:,.2f}"])
    story.append(make_table(['Person','Annual Total','Biz %','Deductible','ITC'], phone_rows,
                            [1.5*inch, 1.2*inch, 0.8*inch, 1.2*inch, 1.2*inch]))

    for person, label in [('greg','Greg'), ('lilibeth','Lilibeth')]:
        months = phone[person].get('months', {})
        if months and any(v > 0 for v in months.values()):
            story.append(Spacer(1, 4))
            story.append(Paragraph(f"{label} Monthly Detail:", styles['Normal']))
            m_rows = [[m, f"${v:,.2f}"] for m, v in months.items() if v > 0]
            story.append(make_table(['Month','Amount'], m_rows, [2*inch, 2*inch]))
    story.append(Spacer(1, 10))

    # 5. GST Summary
    story.append(PageBreak())
    story.append(Paragraph("5. GST/HST Filing Summary", styles['SectionHead']))
    total_itc = total_bank_itc + cash_itc + total_phone_itc
    net_gst = gst_collected - total_itc
    gst_rows = [
        ['Line 101 - Revenue (before GST)', f"${rev['grand_total']:,.2f}"],
        ['Line 105 - GST Collected (5%)', f"${gst_collected:,.2f}"],
        ['', ''],
        ['ITC - Bank Transactions', f"${total_bank_itc:,.2f}"],
        ['ITC - Cash Expenses', f"${cash_itc:,.2f}"],
        ['ITC - Phone Bills', f"${total_phone_itc:,.2f}"],
        ['Line 108 - Total ITCs', f"${total_itc:,.2f}"],
        ['', ''],
    ]
    if net_gst > 0:
        gst_rows.append(['LINE 109 - NET GST OWING', f"${net_gst:,.2f}"])
    else:
        gst_rows.append(['LINE 109 - GST REFUND', f"${abs(net_gst):,.2f}"])
    story.append(make_table(['Item','Amount'], gst_rows, [4*inch, 2*inch]))
    story.append(Spacer(1, 10))

    # 6. Shareholder Split
    story.append(Paragraph("6. Shareholder Income Split", styles['SectionHead']))
    net_income = rev['grand_total'] - total_exp - sum(e.get('amount',0) for e in (cash_expenses or []))
    split_rows = [
        ['Total Revenue', f"${rev['grand_total']:,.2f}"],
        ['Total Expenses', f"${total_exp + sum(e.get('amount',0) for e in (cash_expenses or [])):,.2f}"],
        ['Net Income', f"${net_income:,.2f}"],
        ['', ''],
        ['Greg MacDonald (51%)', f"${net_income * 0.51:,.2f}"],
        ['Lilibeth Sejera (49%)', f"${net_income * 0.49:,.2f}"],
    ]
    story.append(make_table(['Item','Amount'], split_rows, [4*inch, 2*inch]))
    story.append(Spacer(1, 20))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
    story.append(Paragraph("Auto-generated by RigBooks. Verify against source documents.", styles['SmallNote']))

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


def generate_excel(classified_df, cash_expenses, phone_bill, fiscal_year="2024-2025"):
    buf = io.BytesIO()
    wb = Workbook()
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill('solid', fgColor='1a1a5e')
    currency_fmt = '$#,##0.00'
    pct_fmt = '0%'
    bold_f = Font(name='Arial', bold=True, size=10)
    normal_f = Font(name='Arial', size=10)
    thin_border = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
                         top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))

    def style_header(ws, row, ncols):
        for c in range(1, ncols+1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    def style_row(ws, row, ncols, is_total=False):
        for c in range(1, ncols+1):
            cell = ws.cell(row=row, column=c)
            cell.font = bold_f if is_total else normal_f
            cell.border = thin_border

    def auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # Revenue sheet
    ws = wb.active
    ws.title = "Revenue"
    rev = _build_revenue_breakdown(classified_df)
    ws.append(["Cape Bretoner's Oilfield Services Ltd."])
    ws.merge_cells('A1:D1')
    ws['A1'].font = Font(name='Arial', bold=True, size=14)
    ws.append([f"Revenue - FY {fiscal_year}"])
    ws['A2'].font = Font(name='Arial', bold=True, size=11, color='666666')
    ws.append([])
    row = 4
    for c, h in enumerate(['Date','Description','Type','Amount'], 1):
        ws.cell(row=row, column=c, value=h)
    style_header(ws, row, 4)
    row += 1
    for label, df_sub in [('Wire Transfer', rev['wire']), ('Mobile Deposit', rev['mobile']),
                           ('Branch Deposit', rev['branch']), ('E-Transfer', rev['etransfer'])]:
        if df_sub is not None and not df_sub.empty:
            for _, r in df_sub.iterrows():
                d = r.get('date','')
                if hasattr(d, 'strftime'): d = d.strftime('%Y-%m-%d')
                ws.cell(row=row, column=1, value=str(d))
                ws.cell(row=row, column=2, value=str(r['description'])[:60])
                ws.cell(row=row, column=3, value=label)
                ws.cell(row=row, column=4, value=r['credit']).number_format = currency_fmt
                style_row(ws, row, 4)
                row += 1
    row += 1
    ws.cell(row=row, column=3, value="TOTAL REVENUE").font = bold_f
    ws.cell(row=row, column=4, value=rev['grand_total']).number_format = currency_fmt
    ws.cell(row=row, column=4).font = bold_f
    style_row(ws, row, 4, is_total=True)
    row += 1
    ws.cell(row=row, column=3, value="GST Collected (5%)").font = normal_f
    ws.cell(row=row, column=4, value=rev['grand_total'] * 0.05).number_format = currency_fmt
    auto_width(ws)

    # Expenses sheet
    ws2 = wb.create_sheet("Expenses by Category")
    exp = _build_expense_breakdown(classified_df)
    ws2.append(["Business Expenses by CRA Category"])
    ws2['A1'].font = Font(name='Arial', bold=True, size=12)
    ws2.append([])
    row = 3
    for c, h in enumerate(['Category','Transactions','Total','ITC'], 1):
        ws2.cell(row=row, column=c, value=h)
    style_header(ws2, row, 4)
    row += 1
    total_exp = 0
    total_bank_itc = 0
    for cat in sorted(exp.keys()):
        info = exp[cat]
        ws2.cell(row=row, column=1, value=cat)
        ws2.cell(row=row, column=2, value=info['count'])
        ws2.cell(row=row, column=3, value=info['total']).number_format = currency_fmt
        ws2.cell(row=row, column=4, value=info['itc']).number_format = currency_fmt
        style_row(ws2, row, 4)
        total_exp += info['total']
        total_bank_itc += info['itc']
        row += 1
    ws2.cell(row=row, column=1, value="TOTAL").font = bold_f
    ws2.cell(row=row, column=3, value=total_exp).number_format = currency_fmt
    ws2.cell(row=row, column=3).font = bold_f
    ws2.cell(row=row, column=4, value=total_bank_itc).number_format = currency_fmt
    ws2.cell(row=row, column=4).font = bold_f
    style_row(ws2, row, 4, is_total=True)
    auto_width(ws2)

    # Cash Expenses sheet
    ws3 = wb.create_sheet("Cash Expenses")
    ws3.append(["Cash Expenses (Not in Bank Statement)"])
    ws3['A1'].font = Font(name='Arial', bold=True, size=12)
    ws3.append([])
    row = 3
    for c, h in enumerate(['Date','Description','Category','Amount','ITC','Receipt'], 1):
        ws3.cell(row=row, column=c, value=h)
    style_header(ws3, row, 6)
    row += 1
    cash_itc = 0
    for e in (cash_expenses or []):
        amt = e.get('amount', 0)
        itc = amt * 0.05 / 1.05
        cash_itc += itc
        receipt = "Yes" if e.get('has_receipt', False) else ("Recommended" if amt >= 30 else "Not required")
        ws3.cell(row=row, column=1, value=e.get('date','N/A'))
        ws3.cell(row=row, column=2, value=e.get('description','N/A'))
        ws3.cell(row=row, column=3, value=e.get('category','N/A'))
        ws3.cell(row=row, column=4, value=amt).number_format = currency_fmt
        ws3.cell(row=row, column=5, value=itc).number_format = currency_fmt
        ws3.cell(row=row, column=6, value=receipt)
        style_row(ws3, row, 6)
        row += 1
    ws3.cell(row=row, column=3, value="TOTAL").font = bold_f
    ws3.cell(row=row, column=4, value=sum(e.get('amount',0) for e in (cash_expenses or []))).number_format = currency_fmt
    ws3.cell(row=row, column=4).font = bold_f
    ws3.cell(row=row, column=5, value=cash_itc).number_format = currency_fmt
    ws3.cell(row=row, column=5).font = bold_f
    style_row(ws3, row, 6, is_total=True)
    auto_width(ws3)

    # Phone Bills sheet
    ws4 = wb.create_sheet("Phone Bills")
    phone = _get_phone_data(phone_bill)
    ws4.append(["Phone Bill Deductions"])
    ws4['A1'].font = Font(name='Arial', bold=True, size=12)
    ws4.append([])
    row = 3
    for c, h in enumerate(['Person','Annual Total','Business %','Deductible','ITC'], 1):
        ws4.cell(row=row, column=c, value=h)
    style_header(ws4, row, 5)
    row += 1
    total_phone_itc = 0
    for person, label in [('greg','Greg MacDonald'), ('lilibeth','Lilibeth Sejera')]:
        p = phone[person]
        ws4.cell(row=row, column=1, value=label)
        ws4.cell(row=row, column=2, value=p['annual']).number_format = currency_fmt
        ws4.cell(row=row, column=3, value=p['business_pct']/100).number_format = pct_fmt
        ws4.cell(row=row, column=4, value=p['deductible']).number_format = currency_fmt
        ws4.cell(row=row, column=5, value=p['itc']).number_format = currency_fmt
        style_row(ws4, row, 5)
        total_phone_itc += p['itc']
        row += 1
    ws4.cell(row=row, column=1, value="TOTAL").font = bold_f
    ws4.cell(row=row, column=4, value=phone['greg']['deductible']+phone['lilibeth']['deductible']).number_format = currency_fmt
    ws4.cell(row=row, column=4).font = bold_f
    ws4.cell(row=row, column=5, value=total_phone_itc).number_format = currency_fmt
    ws4.cell(row=row, column=5).font = bold_f
    style_row(ws4, row, 5, is_total=True)
    auto_width(ws4)

    # GST Filing sheet
    ws5 = wb.create_sheet("GST Filing")
    gst_collected = rev['grand_total'] * 0.05
    total_itc = total_bank_itc + cash_itc + total_phone_itc
    net_gst = gst_collected - total_itc
    ws5.append(["GST/HST Filing Summary"])
    ws5['A1'].font = Font(name='Arial', bold=True, size=14)
    ws5.append([f"Fiscal Year {fiscal_year}"])
    ws5['A2'].font = Font(name='Arial', size=11, color='666666')
    ws5.append([])
    row = 4
    items = [
        ('Line 101 - Revenue (before GST)', rev['grand_total']),
        ('Line 105 - GST Collected (5%)', gst_collected),
        ('', None),
        ('ITC - Bank Transactions', total_bank_itc),
        ('ITC - Cash Expenses', cash_itc),
        ('ITC - Phone Bills', total_phone_itc),
        ('Line 108 - Total ITCs', total_itc),
        ('', None),
    ]
    items.append(('LINE 109 - NET GST OWING' if net_gst > 0 else 'LINE 109 - GST REFUND', abs(net_gst)))
    for label, val in items:
        is_bold = 'TOTAL' in label or 'LINE 109' in label or 'Line 108' in label
        ws5.cell(row=row, column=1, value=label).font = bold_f if is_bold else normal_f
        if val is not None:
            ws5.cell(row=row, column=2, value=val).number_format = currency_fmt
            if is_bold: ws5.cell(row=row, column=2).font = bold_f
        row += 1
    auto_width(ws5)

    # All Transactions sheet
    if classified_df is not None and not classified_df.empty:
        ws6 = wb.create_sheet("All Transactions")
        cols = [c for c in ['date','description','debit','credit','cra_category','itc_amount'] if c in classified_df.columns]
        ws6.append(cols)
        style_header(ws6, 1, len(cols))
        for _, r in classified_df[cols].iterrows():
            row_data = []
            for c in cols:
                val = r[c]
                if hasattr(val, 'strftime'): val = val.strftime('%Y-%m-%d')
                elif pd.isna(val): val = ''
                row_data.append(val)
            ws6.append(row_data)
        auto_width(ws6)

    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
